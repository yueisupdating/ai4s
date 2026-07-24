#!/usr/bin/env python3
"""Build the AI4S questionnaire summary/chart workbook from score tables.

The demo workbook is treated as read-only.  By default this script reads:

  D:\\coding\\codex\\score_tables\\*.xlsx

and writes:

  D:\\coding\\codex\\问卷图表设计20260715_汇总.xlsx

It copies the demo workbook layout first, then refreshes all data sheets and
chart references in the copied workbook.
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, RadarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


ROOT = Path(__file__).resolve().parent
DEFAULT_SCORE_DIR = ROOT / "score_tables"
DEFAULT_TEMPLATE = ROOT / "问卷图表设计20260715.xlsx"
DEFAULT_OUTPUT = ROOT / "问卷图表设计20260715_汇总.xlsx"

FIELDS = [
    "数学与系统科学",
    "物理",
    "化学",
    "空间科学",
    "生命科学",
    "地球科学",
    "信息科技",
    "材料科学",
    "能源科学",
    "海洋科学",
    "环境与生态",
    "力学",
    "精密仪器与装备",
    "土木水利",
    "大科学工程",
    "交叉科学",
]

SECTION_SCORES = [
    ("数据资源就绪度", "数据资源就绪度"),
    ("计算资源就绪度", "计算资源就绪度"),
    ("模型与平台就绪度", "模型与平台就绪度"),
    ("科研创新就绪度", "科研创新就绪度"),
    ("人才储备就绪度", "人才储备就绪度"),
    ("组织机制就绪度", "组织机制就绪度"),
    ("开放合作就绪度", "开放合作就绪度"),
    ("整体就绪度", "整体就绪度"),
]

TABLE_SIDE = Side(style="thin", color="D9E2F3")
TABLE_BORDER = Border(left=TABLE_SIDE, right=TABLE_SIDE, top=TABLE_SIDE, bottom=TABLE_SIDE)
TABLE_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TABLE_HEADER_FONT = Font(bold=True, color="1F2937")
TABLE_BODY_FILL = PatternFill(fill_type=None)
TABLE_BODY_FONT = Font(color="111827")

DATA_QUESTIONS = [
    "科研数据整体情况",
    "数据管理平台建设情况",
    "科研数据可访问性",
    "内部数据共享机制完善度",
    "外部数据共享合作参与度",
]

COMPUTE_QUESTIONS = [
    "计算资源来源构成情况",
    "AI4S计算资源科研需求满足情况",
    "专用AI计算硬件（如GPU集群）情况",
    "云计算资源便捷使用情况",
    "计算资源分配机制合理性",
]

MODEL_QUESTIONS = [
    "自主研发AI科研工具或平台情况",
    "领域专用AI模型（垂域模型）研发情况",
    "AI支撑平台使用情况",
]

INNOVATION_QUESTIONS = [
    "科研人员AI使用情况",
    "AI4S相关项目数量和质量",
    "AI驱动重大科研成果情况",
    "AI科研成果转化能力",
]

TALENT_QUESTIONS = [
    "复合型人才储备情况",
    "AI相关培训情况",
    "AI4S人才激励机制情况",
]

ORG_QUESTIONS = [
    "科研组织机制调整情况",
    "AI4S跨学科团队建设支持情况",
    "分布式科研团队组建情况",
    "AI4S科研绩效考核适配情况",
    "高风险创新项目支持情况",
    "AI4S相关成果知识产权保护情况",
]

OPEN_QUESTIONS = [
    "AI领域外部合作情况",
    "国际AI4S科研合作情况",
    "AI4S开源社区或开放科学平台参与情况",
    "内部合作文化",
]

READINESS_DETAIL_BLOCKS = [
    ("数据资源就绪度", "数据资源就绪度", DATA_QUESTIONS),
    ("计算资源就绪度", "计算资源就绪度", COMPUTE_QUESTIONS),
    ("模型与平台就绪度", "模型与平台就绪度", MODEL_QUESTIONS),
    ("科研创新就绪度", "科研创新就绪度", INNOVATION_QUESTIONS),
    ("人才储备就绪度", "人才储备就绪度", TALENT_QUESTIONS),
    ("组织机制就绪度", "组织机制就绪度", ORG_QUESTIONS),
    ("开放合作就绪度", "开放合作就绪度", OPEN_QUESTIONS),
]

ALL_NUMERIC_QUESTIONS = (
    DATA_QUESTIONS
    + COMPUTE_QUESTIONS
    + MODEL_QUESTIONS
    + INNOVATION_QUESTIONS
    + TALENT_QUESTIONS
    + ORG_QUESTIONS
    + OPEN_QUESTIONS
    + ["整体就绪度"]
)

Q17_OPTIONS = [
    ("B", "知识管理与文献研究"),
    ("C", "海量数据采集与预处理"),
    ("E", "复杂现象模拟与预测"),
    ("G", "研究成果分析与验证"),
    ("D", "实验方案设计与优化"),
    ("F", "未知规律挖掘与理论推导"),
    ("A", "暂未应用"),
    ("H", "其他"),
]

Q35_OPTIONS = [
    ("B", "提升效率、缩短研究周期"),
    ("D", "突破传统方法、解决复杂问题"),
    ("C", "优化实验设计与模拟"),
    ("G", "促进跨学科协同创新"),
    ("E", "产出创新性成果"),
    ("F", "提高研究成果准确性与可靠性"),
    ("A", "感觉价值不大"),
    ("H", "其他"),
]

Q36_OPTIONS = [
    ("F", "缺乏复合型人才"),
    ("D", "领域模型、平台支撑不足"),
    ("B", "数据匮乏、共享困难"),
    ("E", "跨学科协同机制不畅"),
    ("G", "考核激励机制不匹配"),
    ("C", "计算资源分配不合理"),
    ("I", "对外合作不足"),
    ("H", "标准规范与伦理机制不足"),
    ("A", "暂未发现使用瓶颈"),
    ("J", "其他"),
]

Q_PLATFORM_USAGE_OPTIONS = [
    ("PANSHI", "使用“磐石Science One ”AI平台"),
    ("CSTCLOUD", "使用中国科技云大模型开放服务平台"),
    ("PUBSCHOLAR", "使用PubScholar公益学术平台"),
]

Q_PLATFORM_USAGE_COUNTS = Counter({"PANSHI": 4, "CSTCLOUD": 2, "PUBSCHOLAR": 5})

Q_ORG_ADJUSTMENT_OPTIONS = [
    ("A", "暂无相关调整"),
    ("B", "优化科研项目立项、选题评审机制，强化 AI 与科学交叉导向"),
    ("C", "组建AI与领域科学融合的跨学科研究团队"),
    ("D", "优化算力、数据等科研资源配置与共享机制"),
    ("E", "改革科研评价与激励机制，适配交叉创新导向"),
    ("F", "建立AI4S科研成果知识产权保护机制"),
    ("G", "建立跨团队、跨部门协同攻关与常态化交流机制"),
    ("H", "建立AI4S复合型科研人才培养机制"),
    ("I", "其他"),
]

Q_DISTRIBUTED_TEAM_OPTIONS = [
    ("1", "不了解分布式团队，未组建任何相关团队"),
    ("2", "仅设有临时项目、课题组科研团队，无固定分布式协作模式"),
    ("3", "组建所内跨实验室、跨所联合团队"),
    ("4", "搭建院内外协同攻关团队，形成成熟的分布式科研体系"),
]

Q38_OPTIONS = [
    ("A", "完全不紧迫"),
    ("B", "不太紧迫"),
    ("C", "比较紧迫"),
    ("D", "极为紧迫"),
]

Q39_OPTIONS = [
    ("A", "无需特别改革"),
    ("B", "建立统一、高质量、可共享的领域数据集"),
    ("C", "建设或完善统一的数据管理平台"),
    ("D", "优化数据访问流程，提升可访问性与自动化调用能力"),
    ("E", "完善内部数据共享机制，打破数据孤岛"),
    ("F", "推动外部数据共享合作，融入行业数据生态"),
    ("G", "其他"),
]

Q40_OPTIONS = [
    ("A", "无需特别改革"),
    ("B", "加大专用AI计算硬件投入与常态化更新"),
    ("C", "优化计算资源分配机制，提高公平性与调度效率"),
    ("D", "降低云计算等外部资源使用门槛，实现灵活调用"),
    ("E", "建设统一算力调度平台，支撑大规模AI科研任务"),
    ("F", "其他"),
]

Q41_OPTIONS = [
    ("A", "暂无相关需求"),
    ("B", "专属算力集群"),
    ("C", "领域数据集"),
    ("D", "垂域模型训练"),
    ("E", "数据安全保障"),
    ("F", "智能科研工具"),
    ("G", "其他"),
]

Q42_OPTIONS = [
    ("A", "无需特别改革优化"),
    ("B", "建立更加扁平化、灵活的组织架构"),
    ("C", "设立跨学科、跨部门、跨机构的协同创新分布式团队"),
    ("D", "成立AI4S专项实验室，设立AI4S专项管理岗位"),
    ("E", "优化算力、数据等科研资源的配置与共享机制"),
    ("F", "改革科研评价与激励机制，适配交叉创新导向"),
    ("G", "建立AI4S科研成果知识产权保护与归属机制"),
    ("H", "完善AI4S复合型人才引进与培养机制"),
    ("I", "推动与AI领域领先机构、产业界的深度合作"),
    ("J", "其他"),
]

Q43_OPTIONS = [
    ("A", "无需任何改革"),
    ("B", "弱化传统论文、项目数量考核，强化AI科研创新价值"),
    ("C", "建立AI科研成果转化、技术应用成效考核指标"),
    ("D", "针对分布式团队、交叉团队制定差异化考核标准"),
    ("E", "增加AI技术融合科研工作的考核权重"),
    ("F", "其他"),
]

Q44_OPTIONS = [
    ("A", "无需任何改革"),
    ("B", "出台细化制度，明确AI生成科研成果的归属规则"),
    ("C", "建立AI科研知识产权快速申报与保护机制"),
    ("D", "完善跨团队合作成果知识产权分配规则"),
    ("E", "加强知识产权相关培训"),
    ("F", "其他"),
]

Q45_OPTIONS = [
    ("A", "无需任何改革"),
    ("B", "提供系统性、常态化的AI+科研交叉培训"),
    ("C", "设立AI4S复合型人才引进专项通道"),
    ("D", "建立有利于AI4S人才的激励机制"),
    ("E", "支持青年科研人员开展AI4S前沿探索"),
    ("F", "其他"),
]

Q46_OPTIONS = [
    ("A", "无需特别推进"),
    ("B", "加强与AI领域领先科研院所或产业界的深度合作"),
    ("C", "积极参与国际AI4S科研合作项目"),
    ("D", "主动贡献或参与AI4S开源社区、开放科学平台"),
    ("E", "培育内部开放协作的研究文化，打破团队壁垒"),
    ("F", "其他"),
]

Q47_DIMENSIONS = [
    ("A", "建立跨学科分布式团队"),
    ("B", "改革考核评价机制"),
    ("C", "明确知识产权归属"),
    ("D", "优化资源配置"),
    ("E", "引进与培养复合型人才"),
    ("F", "培育开放合作与协同文化"),
]

RANK_TO_COLUMN = {
    "第一": "第一优先级",
    "第二": "第二优先级",
    "第三": "第三优先级",
    "第四": "第四优先级",
}


@dataclass
class InstituteRecord:
    name: str
    fields: list[str]
    q47_answer: str
    values_by_label: dict[str, list[object]]


INSTITUTE_DISPLAY_NAMES = {
    "中科院动物所": "中国科学院动物研究所",
    "中科院植物所": "中国科学院植物研究所",
    "中科院上海药物所": "中国科学院上海药物研究所",
    "中科院地质与地球物理所": "中国科学院地质与地球物理研究所",
    "中科院大气物理所": "中国科学院大气物理研究所",
    "中科院半导体所": "中国科学院半导体研究所",
    "中科院微电子所": "中国科学院微电子研究所",
    "中科院数学与系统科学研究院": "中国科学院数学与系统科学研究院",
    "中科院物理所": "中国科学院物理研究所",
    "中科院化学研究所": "中国科学院化学研究所",
    "中科院化学所": "中国科学院化学研究所",
    "中科院国家空间科学中心": "中国科学院国家空间科学中心",
    "海洋所": "中国科学院海洋研究所",
    "武汉岩土力学所": "中国科学院武汉岩土力学研究所",
    "工程热物理所": "中国科学院工程热物理研究所",
    "宁波材料技术与工程所": "中国科学院宁波材料技术与工程研究所",
    "合肥物质科学研究院": "中国科学院合肥物质科学研究院",
    "力学所": "中国科学院力学研究所",
    "空天信息创新研究院": "中国科学院空天信息创新研究院",
    "生态环境研究中心": "中国科学院生态环境研究中心",
    "深圳先进技术研究院": "中国科学院深圳先进技术研究院",
    "长春光机所": "中国科学院长春光学精密机械与物理研究所",
    "高能物理所": "中国科学院高能物理研究所",
    "金属研究所": "中国科学院金属研究所",
    "东北地理与农业生态研究所": "中国科学院东北地理与农业生态研究所",
    "上海技术物理研究所": "中国科学院上海技术物理研究所",
    "分子细胞科学卓越创新中心": "中国科学院分子细胞科学卓越创新中心",
    "上海营养与健康研究所": "中国科学院上海营养与健康研究所",
    "福建物质结构研究所": "中国科学院福建物质结构研究所",
    "南京地质古生物研究所": "中国科学院南京地质古生物研究所",
    "紫金山天文台": "中国科学院紫金山天文台",
    "苏州生物医学工程技术研究所": "中国科学院苏州生物医学工程技术研究所",
    "水生生物研究所": "中国科学院水生生物研究所",
    "南海海洋研究所": "中国科学院南海海洋研究所",
    "广州地球化学研究所": "中国科学院广州地球化学研究所",
    "东莞材料科学与技术研究所": "中国科学院东莞材料科学与技术研究所",
    "成都生物研究所": "中国科学院成都生物研究所",
    "重庆绿色智能技术研究院": "中国科学院重庆绿色智能技术研究院",
    "西双版纳热带植物园": "中国科学院西双版纳热带植物园",
    "国家授时中心": "中国科学院国家授时中心",
    "近代物理研究所": "中国科学院近代物理研究所",
    "青海盐湖研究所": "中国科学院青海盐湖研究所",
    "新疆生态与地理研究所": "中国科学院新疆生态与地理研究所",
    "古脊椎动物与古人类研究所": "中国科学院古脊椎动物与古人类研究所",
    "生物物理研究所": "中国科学院生物物理研究所",
    "北京基因组研究所（国家生物信息中心）": "中国科学院北京基因组研究所（国家生物信息中心）",
    "自动化研究所": "中国科学院自动化研究所",
    "科技战略咨询研究院": "中国科学院科技战略咨询研究院",
    "空间应用工程与技术中心": "中国科学院空间应用工程与技术中心",
}


def formal_institute_name(name: str) -> str:
    name = name.strip()
    if name in INSTITUTE_DISPLAY_NAMES:
        return INSTITUTE_DISPLAY_NAMES[name]
    if name.startswith("中国科学院"):
        return name
    if name.startswith("中科院"):
        remainder = name.removeprefix("中科院")
        return INSTITUTE_DISPLAY_NAMES.get(remainder, f"中国科学院{remainder}")
    if name.endswith(("研究所", "研究院", "中心", "天文台", "植物园")):
        return f"中国科学院{name}"
    return name


def clean_institute_name(path: Path) -> str:
    stem = path.stem
    return formal_institute_name(stem.removeprefix("评分表_"))


def normalize_number(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if re.fullmatch(r"-?\d+(?:\.\d+)?", stripped):
            return float(stripped)
    return None


def rounded(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 2)


def avg(values: Iterable[object]) -> float | None:
    nums = [number for value in values if (number := normalize_number(value)) is not None]
    if not nums:
        return None
    return rounded(mean(nums))


def split_fields(value: object) -> list[str]:
    if not isinstance(value, str):
        return []
    fields: list[str] = []
    for item in re.split(r"[、,，;；]\s*", value):
        item = item.strip()
        if item in FIELDS and item not in fields:
            fields.append(item)
    return fields


def selected_letters(value: object) -> list[str]:
    if not isinstance(value, str):
        return []
    return re.findall(r"[A-Z]", value.upper())


def first_value(record: InstituteRecord, label: str) -> object:
    values = record.values_by_label.get(label, [])
    return values[0] if values else None


def last_value(record: InstituteRecord, label: str) -> object:
    values = record.values_by_label.get(label, [])
    return values[-1] if values else None


def numeric_score(record: InstituteRecord, label: str, last: bool = False) -> float | None:
    value = last_value(record, label) if last else first_value(record, label)
    return normalize_number(value)


def read_score_table(path: Path) -> InstituteRecord:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active

    values_by_label: dict[str, list[object]] = defaultdict(list)
    for row in worksheet.iter_rows(min_col=1, max_col=2, values_only=True):
        label, value = row
        if label is None:
            continue
        values_by_label[str(label)].append(value)

    fields = split_fields(first_value_from_dict(values_by_label, "主要研究领域（至多三项）"))
    q47_answer = str(first_value_from_dict(values_by_label, "最后一个问题回答（Q47）") or "").strip()
    return InstituteRecord(
        name=clean_institute_name(path),
        fields=fields,
        q47_answer=q47_answer,
        values_by_label=values_by_label,
    )


def first_value_from_dict(values_by_label: dict[str, list[object]], label: str) -> object:
    values = values_by_label.get(label, [])
    return values[0] if values else None


def load_records(score_dir: Path) -> list[InstituteRecord]:
    paths = sorted(score_dir.glob("评分表_*.xlsx"))
    if not paths:
        raise SystemExit(f"未找到评分表：{score_dir}")
    return [read_score_table(path) for path in paths]


def clear_range(ws, min_row: int, max_row: int, min_col: int, max_col: int, clear_format: bool = False) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None
            if clear_format:
                cell.font = Font()
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                cell.alignment = Alignment()
                cell.number_format = "General"


def write_table(ws, start_row: int, start_col: int, rows: Sequence[Sequence[object]]) -> None:
    for row_offset, row in enumerate(rows):
        for col_offset, value in enumerate(row):
            cell = ws.cell(start_row + row_offset, start_col + col_offset)
            cell.value = value
            cell.border = TABLE_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if row_offset == 0:
                cell.fill = TABLE_HEADER_FILL
                cell.font = TABLE_HEADER_FONT
            else:
                cell.fill = TABLE_BODY_FILL
                cell.font = TABLE_BODY_FONT
                if isinstance(value, str) and "\n" in value:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")


def percent(count: int, total: int) -> float:
    return round(count / total, 4) if total else 0


def count_letters(records: Sequence[InstituteRecord], label: str) -> Counter[str]:
    counts: Counter[str] = Counter()
    for record in records:
        counts.update(set(selected_letters(first_value(record, label))))
    return counts


def count_fields(records: Sequence[InstituteRecord]) -> Counter[str]:
    counts: Counter[str] = Counter()
    for record in records:
        counts.update(record.fields)
    return counts


def count_scores(records: Sequence[InstituteRecord], label: str) -> Counter[str]:
    counts: Counter[str] = Counter()
    for record in records:
        score = normalize_number(first_value(record, label))
        if score is not None:
            counts[str(int(score))] += 1
    return counts


def q38_letter(record: InstituteRecord) -> str | None:
    value = first_value(record, "AI时代科研组织模式与资源配置改革紧迫性")
    if not isinstance(value, str):
        return None
    match = re.search(r"[A-D]", value.upper())
    return match.group(0) if match else None


def q47_rank_counts(records: Sequence[InstituteRecord]) -> dict[str, Counter[str]]:
    result = {rank: Counter() for rank in RANK_TO_COLUMN}
    rank_pattern = re.compile(r"^(第一|第二|第三|第四)\s*[：:]\s*(.+)$")
    dimensions = [dimension for _, dimension in Q47_DIMENSIONS]

    for record in records:
        for raw_line in record.q47_answer.splitlines():
            line = raw_line.strip()
            match = rank_pattern.match(line)
            if not match:
                continue
            rank, body = match.groups()
            selected = next((dimension for dimension in dimensions if dimension in body), None)
            if selected:
                result[rank][selected] += 1
    return result


def records_for_field(records: Sequence[InstituteRecord], field: str) -> list[InstituteRecord]:
    return [record for record in records if field in record.fields]


def average_records(records: Sequence[InstituteRecord], label: str, last: bool = False) -> float | None:
    values = [numeric_score(record, label, last=last) for record in records]
    return avg(value for value in values if value is not None)


def populate_field_distribution(wb, records: Sequence[InstituteRecord]) -> None:
    ws = wb["研究领域分布"]
    ws._charts = []
    clear_range(ws, 1, 40, 1, 4)

    counts = Counter(field for record in records for field in record.fields)
    rows = [["研究领域", "选中数量"]]
    rows.extend([field, counts[field]] for field in FIELDS)
    write_table(ws, 1, 1, rows)

    chart = BarChart()
    chart.type = "bar"
    chart.title = "研究领域分布"
    chart.y_axis.title = "研究领域"
    chart.x_axis.title = "选中数量"
    chart.height = 10
    chart.width = 18
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=len(rows)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(rows)))
    ws.add_chart(chart, "D2")


def populate_overall_readiness(wb, records: Sequence[InstituteRecord]) -> None:
    ws = wb["就绪度-整体"]
    ws._charts = []
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    clear_range(ws, 1, max(ws.max_row, 180), 1, max(ws.max_column, 18), clear_format=True)

    section_rows = [["一级维度名称", "平均得分（最高分4分）"]]
    for display, label in SECTION_SCORES:
        section_rows.append([display, average_records(records, label)])
    write_table(ws, 1, 1, section_rows)

    add_radar_chart(ws, "AI赋能科研就绪度总览", 1, 1, 9, 2, "E1")

    blocks = [
        ("数据资源就绪度", DATA_QUESTIONS),
        ("计算资源就绪度", COMPUTE_QUESTIONS),
        ("模型与平台就绪度", MODEL_QUESTIONS),
        ("科研创新就绪度", INNOVATION_QUESTIONS),
        ("人才储备就绪度", TALENT_QUESTIONS),
        ("组织机制就绪度", ORG_QUESTIONS),
        ("开放合作就绪度", OPEN_QUESTIONS),
    ]
    start_row = 14
    for title, questions in blocks:
        row_count = write_question_block(ws, start_row, 1, questions, records, title)
        add_bar_chart(ws, title, start_row, 1, start_row + row_count - 1, 2, f"E{start_row}")
        start_row += max(row_count + 3, 12)
    ws.print_area = f"A1:K{start_row - 1}"


def write_question_block(
    ws,
    start_row: int,
    start_col: int,
    questions: Sequence[str],
    records: Sequence[InstituteRecord],
    header: str = "二级题目",
) -> int:
    rows = [[header, "平均得分（最高4分）"]]
    rows.extend([question, average_records(records, question, last=(question == "整体就绪度"))] for question in questions)
    write_table(ws, start_row, start_col, rows)
    return len(rows)


def populate_field_readiness(wb, records: Sequence[InstituteRecord]) -> None:
    ws = wb["就绪度-领域整体"]
    ws._charts = []
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    clear_range(ws, 1, max(ws.max_row, 217), 1, max(ws.max_column, 45), clear_format=True)

    section_rows: list[list[object]] = [["研究领域", *FIELDS]]
    for display, label in SECTION_SCORES:
        section_rows.append([display, *[average_records(records_for_field(records, field), label) for field in FIELDS]])
    write_table(ws, 1, 1, section_rows)

    add_field_radar_chart(ws, "院内各领域", 1, 1, 9, 17, "C11")

    detail_start_row = 1
    detail_start_col = 21
    detail_end_col = detail_start_col + len(FIELDS)
    for display, label, questions in READINESS_DETAIL_BLOCKS:
        rows: list[list[object]] = [[display, *FIELDS]]
        for question in questions:
            rows.append(
                [question, *[average_records(records_for_field(records, field), question) for field in FIELDS]]
            )
        rows.append(
            [
                f"{display}总得分",
                *[average_records(records_for_field(records, field), label) for field in FIELDS],
            ]
        )
        write_table(ws, detail_start_row, detail_start_col, rows)
        add_field_radar_chart(
            ws,
            f"院内各领域{display}",
            detail_start_row,
            detail_start_col,
            detail_start_row + len(rows) - 1,
            detail_end_col,
            f"R{detail_start_row + 13}",
        )
        detail_start_row += 31
    ws.print_area = f"A1:AK{detail_start_row - 1}"


def write_institute_detail_block(
    ws,
    start_row: int,
    start_col: int,
    field: str,
    records: Sequence[InstituteRecord],
    questions: Sequence[str],
) -> int:
    rows: list[list[object]] = [["研究领域", "研究机构", *questions]]
    for record in records:
        rows.append([field, record.name, *[numeric_score(record, question) for question in questions]])
    write_table(ws, start_row, start_col, rows)
    for row in range(start_row + 1, start_row + len(rows)):
        for col in range(start_col + 2, start_col + 2 + len(questions)):
            ws.cell(row, col).number_format = "0.00"
    return len(rows)


def populate_institute_readiness(wb, records: Sequence[InstituteRecord]) -> None:
    ws = wb["就绪度-领域内研究所"]
    ws._charts = []
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    clear_range(ws, 1, max(ws.max_row, 320), 1, max(ws.max_column, 42), clear_format=True)

    section_header = ["研究领域", "研究机构", *[display for display, _ in SECTION_SCORES]]
    start_row = 1
    for field in FIELDS:
        field_records = records_for_field(records, field)
        if not field_records:
            write_table(ws, start_row, 1, [[field, "无对应研究所"]])
            start_row += 3
            continue

        section_rows = [section_header]
        for record in field_records:
            section_rows.append(
                [field, record.name, *[numeric_score(record, label) for _, label in SECTION_SCORES]]
            )
        write_table(ws, start_row, 1, section_rows)

        last_row = start_row + len(field_records)
        add_institute_radar_chart(
            ws,
            f"{field}领域内研究所就绪度",
            start_row,
            2,
            last_row,
            len(section_header),
            f"L{start_row}",
        )
        start_row += max(len(section_rows) + 2, 18)

    detail_start_col = 19
    detail_chart_col = detail_start_col + max(len(questions) for _, _, questions in READINESS_DETAIL_BLOCKS) + 3
    detail_chart_col_letter = get_column_letter(detail_chart_col)
    detail_row = 1
    for field in FIELDS:
        field_records = records_for_field(records, field)
        if not field_records:
            write_table(ws, detail_row, detail_start_col, [[field, "无对应研究所"]])
            detail_row += 4
            continue

        for display, _label, questions in READINESS_DETAIL_BLOCKS:
            row_count = write_institute_detail_block(
                ws,
                detail_row,
                detail_start_col,
                field,
                field_records,
                questions,
            )
            last_row = detail_row + row_count - 1
            add_institute_detail_radar_chart(
                ws,
                f"{field}领域内研究所{display}",
                detail_row,
                detail_start_col + 1,
                last_row,
                detail_start_col + 1 + len(questions),
                f"{detail_chart_col_letter}{detail_row}",
            )
            detail_row += max(row_count + 2, 16)


def populate_multiselect(wb, records: Sequence[InstituteRecord]) -> None:
    ws = wb["就绪度-多选"]
    ws._charts = []
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    clear_range(ws, 1, max(ws.max_row, 260), 1, max(ws.max_column, 16), clear_format=True)
    total = len(records)

    blocks = [
        ("主要研究领域（至多三项）", [(field, field) for field in FIELDS], count_fields(records), "选中数量"),
        ("AI支撑平台使用情况（选项分布）", Q_PLATFORM_USAGE_OPTIONS, Q_PLATFORM_USAGE_COUNTS, "选中数量"),
        ("AI主要应用环节", Q17_OPTIONS, count_letters(records, "AI主要应用环节"), "选中数量"),
        ("科研组织机制调整内容", Q_ORG_ADJUSTMENT_OPTIONS, count_letters(records, "科研组织机制调整内容"), "选中数量"),
        (
            "分布式科研团队组建情况（选项分布）",
            Q_DISTRIBUTED_TEAM_OPTIONS,
            count_scores(records, "分布式科研团队组建情况"),
            "机构数量",
            "其他（如有相同选择可列出，没有不统计）",
        ),
        ("AI带来的核心价值", Q35_OPTIONS, count_letters(records, "AI带来的核心价值"), "选中数量"),
        ("AI科研应用发展核心瓶颈", Q36_OPTIONS, count_letters(records, "AI科研应用发展核心瓶颈"), "选中数量"),
    ]
    last_row = write_choice_blocks(ws, blocks, total, chart_mode="combo")
    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)
    ws.print_area = f"A1:M{last_row}"


def write_choice_blocks(
    ws,
    blocks: Sequence[tuple],
    total: int,
    combo_chart: bool = False,
    chart_mode: str = "count_bar",
) -> int:
    start_row = 1
    for block in blocks:
        title, options, counts, count_header = block[:4]
        note = block[4] if len(block) > 4 else None
        use_combo = combo_chart or chart_mode == "combo"
        row_count = write_option_block(
            ws,
            start_row,
            1,
            title,
            options,
            counts,
            total,
            count_header=count_header,
            include_chart_labels=use_combo,
        )
        if chart_mode == "percent_bar":
            add_percent_bar_chart(ws, title, start_row, 1, start_row + row_count - 1, f"E{start_row}")
        elif use_combo:
            add_combo_chart(ws, title, start_row, 4, start_row + row_count - 1, f"E{start_row}")
        else:
            add_bar_chart(ws, title, start_row, 1, start_row + row_count - 1, 2, f"E{start_row}")
        if note:
            note_row = start_row + row_count
            note_cols = 4 if use_combo else 3
            for col in range(1, note_cols + 1):
                cell = ws.cell(note_row, col)
                cell.value = note if col == 1 else None
                cell.border = TABLE_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.fill = TABLE_BODY_FILL
                cell.font = TABLE_BODY_FONT
            row_count += 1
        start_row += max(row_count + 3, 18)
    return start_row - 1


def write_option_block(
    ws,
    start_row: int,
    start_col: int,
    header: str,
    options: Sequence[tuple[str, str]],
    counts: Counter[str],
    total: int,
    count_header: str = "选中数量",
    include_chart_labels: bool = False,
) -> int:
    rows = [[header, count_header, "占比（=数量/问卷数量）"]]
    if include_chart_labels:
        rows[0].append("编号")
    for idx, (key, label) in enumerate(options, 1):
        row = [label, counts[key], percent(counts[key], total)]
        if include_chart_labels:
            row.append(str(idx))
        rows.append(row)
    write_table(ws, start_row, start_col, rows)
    percent_col = start_col + 2
    for row in range(start_row + 1, start_row + len(rows)):
        ws.cell(row, percent_col).number_format = "0.0%"
    return len(rows)


def write_positive_pie_source(
    ws,
    source_start_row: int,
    source_start_col: int,
    options: Sequence[tuple[str, str]],
    counts: Counter[str],
    target_start_row: int,
    target_start_col: int,
) -> int:
    rows = [["紧迫程度", "选中数量"]]
    for idx, (key, _label) in enumerate(options, 1):
        if counts[key] <= 0:
            continue
        source_row = source_start_row + idx
        rows.append(
            [
                f"={ws.cell(source_row, source_start_col).coordinate}",
                f"={ws.cell(source_row, source_start_col + 1).coordinate}",
            ]
        )
    if len(rows) == 1:
        for idx, (_key, _label) in enumerate(options, 1):
            source_row = source_start_row + idx
            rows.append(
                [
                    f"={ws.cell(source_row, source_start_col).coordinate}",
                    f"={ws.cell(source_row, source_start_col + 1).coordinate}",
                ]
            )

    write_table(ws, target_start_row, target_start_col, rows)
    return len(rows)


def populate_demand(wb, records: Sequence[InstituteRecord]) -> None:
    ws = wb["需求"]
    ws._charts = []
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    clear_range(ws, 1, max(ws.max_row, 220), 1, max(ws.max_column, 20), clear_format=True)
    total = len(records)

    urgency_counts = Counter(letter for record in records if (letter := q38_letter(record)))
    urgency_row_count = write_option_block(
        ws,
        1,
        1,
        "科研组织模式与资源配置改革紧迫性",
        Q38_OPTIONS,
        urgency_counts,
        total,
    )
    pie_source_row = 1
    pie_source_col = 18
    pie_row_count = write_positive_pie_source(
        ws,
        1,
        1,
        Q38_OPTIONS,
        urgency_counts,
        pie_source_row,
        pie_source_col,
    )
    add_pie_chart(ws, "科研组织模式与资源配置改革紧迫性", pie_source_row, pie_source_col, pie_source_row + pie_row_count - 1, pie_source_col + 1, "E1")

    demand_blocks = [
        ("数据资源改革需求", Q39_OPTIONS, count_letters(records, "数据资源改革需求"), "选中数量"),
        ("计算资源改革需求", Q40_OPTIONS, count_letters(records, "计算资源改革需求"), "选中数量"),
        ("专属AI平台建设需求", Q41_OPTIONS, count_letters(records, "专属AI平台建设需求"), "选中数量"),
        ("组织机制优化方向", Q42_OPTIONS, count_letters(records, "组织机制优化方向"), "选中数量"),
        ("考核评价机制改革需求", Q43_OPTIONS, count_letters(records, "考核评价机制改革需求"), "选中数量"),
        ("知识产权归属与保护改革需求", Q44_OPTIONS, count_letters(records, "知识产权归属与保护改革需求"), "选中数量"),
        ("AI4S人才培养与引进改革需求", Q45_OPTIONS, count_letters(records, "AI4S人才培养与引进改革需求"), "选中数量"),
        ("开放合作推进需求", Q46_OPTIONS, count_letters(records, "开放合作推进需求"), "选中数量"),
    ]
    start_row = 19
    for title, options, counts, count_header in demand_blocks:
        row_count = write_option_block(
            ws,
            start_row,
            1,
            title,
            options,
            counts,
            total,
            count_header=count_header,
            include_chart_labels=True,
        )
        add_combo_chart(ws, title, start_row, 4, start_row + row_count - 1, f"E{start_row}")
        start_row += max(row_count + 3, 18)

    q47_start_row = start_row + 1
    rank_counts = q47_rank_counts(records)
    q47_rows = [["诉求维度", "第一优先级", "第二优先级", "第三优先级", "第四优先级", "合计频次"]]
    ranked_dimensions = []
    for letter, dimension in Q47_DIMENSIONS:
        rank_values = [rank_counts[rank][dimension] for rank in RANK_TO_COLUMN]
        total_count = sum(rank_values)
        ranked_dimensions.append((letter, dimension, rank_values, total_count))
    for letter, dimension, rank_values, total_count in sorted(
        ranked_dimensions,
        key=lambda item: (
            -item[3],
            -item[2][0],
            -item[2][1],
            item[0],
        ),
    ):
        q47_rows.append([f"{letter} {dimension}", *rank_values, total_count])
    write_table(ws, q47_start_row, 1, q47_rows)
    add_q47_chart(
        ws,
        "核心科研组织管理创新诉求",
        q47_start_row,
        1,
        q47_start_row + len(q47_rows) - 1,
        f"H{q47_start_row}",
    )
    ws.print_area = f"A1:Q{q47_start_row + max(len(q47_rows), 16)}"
    ws.row_breaks.append(Break(id=54))
    ws.row_breaks.append(Break(id=108))
    ws.row_breaks.append(Break(id=144))


def add_bar_chart(ws, title: str, header_row: int, category_col: int, max_row: int, value_col: int, anchor: str) -> None:
    chart = BarChart()
    chart.type = "bar"
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.height = 7.5
    chart.width = 12
    chart.add_data(Reference(ws, min_col=value_col, min_row=header_row, max_row=max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=category_col, min_row=header_row + 1, max_row=max_row))
    ws.add_chart(chart, anchor)


def add_percent_bar_chart(ws, title: str, header_row: int, category_col: int, max_row: int, anchor: str) -> None:
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = title
    option_count = max(1, max_row - header_row)
    chart.style = 2
    chart.height = min(max(6.5, option_count * 0.45 + 2.6), 11)
    chart.width = 17.5
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.x_axis.numFmt = "0.0%"
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 1.12
    chart.add_data(Reference(ws, min_col=3, min_row=header_row, max_row=max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=category_col, min_row=header_row + 1, max_row=max_row))
    if chart.series:
        chart.series[0].tx = SeriesLabel(v="占比")
        chart.series[0].graphicalProperties.solidFill = "5B9BD5"
        chart.series[0].graphicalProperties.line.solidFill = "5B9BD5"
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showLegendKey = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.showPercent = False
    chart.dLbls.showBubbleSize = False
    chart.dLbls.dLblPos = "outEnd"
    chart.dLbls.numFmt = "0.0%"
    chart.legend = None
    ws.add_chart(chart, anchor)


def add_combo_chart(ws, title: str, header_row: int, category_col: int, max_row: int, anchor: str) -> None:
    count_chart = BarChart()
    count_chart.type = "col"
    count_chart.title = title
    count_chart.style = 2
    count_chart.height = 7.6
    count_chart.width = 14.5
    count_chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    count_chart.y_axis.title = None
    count_chart.y_axis.numFmt = "0"
    count_chart.y_axis.majorGridlines = None
    count_chart.x_axis.title = "选项编号"
    count_chart.overlap = -27
    count_chart.add_data(Reference(ws, min_col=2, min_row=header_row, max_row=max_row), titles_from_data=True)
    count_chart.set_categories(Reference(ws, min_col=category_col, min_row=header_row + 1, max_row=max_row))
    if count_chart.series:
        count_chart.series[0].tx = SeriesLabel(v="数量")
        count_chart.series[0].graphicalProperties.solidFill = "5B9BD5"
        count_chart.series[0].graphicalProperties.line.solidFill = "5B9BD5"

    percent_chart = LineChart()
    percent_chart.style = 2
    percent_chart.y_axis.axId = 200
    percent_chart.y_axis.title = None
    percent_chart.y_axis.numFmt = "0.0%"
    percent_chart.y_axis.scaling.min = 0
    percent_chart.y_axis.scaling.max = 1
    percent_chart.y_axis.crosses = "max"
    percent_chart.add_data(Reference(ws, min_col=3, min_row=header_row, max_row=max_row), titles_from_data=True)
    percent_chart.set_categories(Reference(ws, min_col=category_col, min_row=header_row + 1, max_row=max_row))
    if percent_chart.series:
        percent_chart.series[0].tx = SeriesLabel(v="占比")
        percent_chart.series[0].graphicalProperties.line.solidFill = "ED7D31"
        percent_chart.series[0].graphicalProperties.line.width = 22000
        percent_chart.series[0].marker.symbol = "circle"
        percent_chart.series[0].marker.size = 5
    percent_chart.dLbls = DataLabelList()
    percent_chart.dLbls.showVal = True
    percent_chart.dLbls.showLegendKey = False
    percent_chart.dLbls.showCatName = False
    percent_chart.dLbls.showSerName = False
    percent_chart.dLbls.showPercent = False
    percent_chart.dLbls.showBubbleSize = False
    percent_chart.dLbls.dLblPos = "r"
    percent_chart.dLbls.numFmt = "0.0%"

    count_chart += percent_chart
    if count_chart.legend:
        count_chart.legend.position = "b"
    ws.add_chart(count_chart, anchor)


def add_radar_chart(ws, title: str, header_row: int, category_col: int, max_row: int, value_col: int, anchor: str) -> None:
    chart = RadarChart()
    chart.type = "standard"
    chart.title = title
    chart.height = 7.5
    chart.width = 10
    chart.add_data(Reference(ws, min_col=value_col, min_row=header_row, max_row=max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=category_col, min_row=header_row + 1, max_row=max_row))
    ws.add_chart(chart, anchor)


def add_field_radar_chart(ws, title: str, header_row: int, start_col: int, max_row: int, max_col: int, anchor: str) -> None:
    chart = RadarChart()
    chart.type = "standard"
    chart.title = title
    chart.height = 8
    chart.width = 14
    chart.add_data(Reference(ws, min_col=start_col + 1, max_col=max_col, min_row=header_row, max_row=max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=start_col, min_row=header_row + 1, max_row=max_row))
    ws.add_chart(chart, anchor)


def add_institute_radar_chart(ws, title: str, header_row: int, start_col: int, max_row: int, max_col: int, anchor: str) -> None:
    chart = RadarChart()
    chart.type = "standard"
    chart.title = title
    chart.height = 8.5
    chart.width = 18
    chart.add_data(
        Reference(ws, min_col=start_col, max_col=max_col, min_row=header_row + 1, max_row=max_row),
        titles_from_data=True,
        from_rows=True,
    )
    chart.set_categories(Reference(ws, min_col=start_col + 1, max_col=max_col, min_row=header_row))
    ws.add_chart(chart, anchor)


def add_institute_detail_radar_chart(ws, title: str, header_row: int, start_col: int, max_row: int, max_col: int, anchor: str) -> None:
    chart = RadarChart()
    chart.type = "standard"
    chart.title = title
    chart.height = 7.2
    chart.width = 13.5
    chart.add_data(
        Reference(ws, min_col=start_col, max_col=max_col, min_row=header_row + 1, max_row=max_row),
        titles_from_data=True,
        from_rows=True,
    )
    chart.set_categories(Reference(ws, min_col=start_col + 1, max_col=max_col, min_row=header_row))
    ws.add_chart(chart, anchor)


def add_pie_chart(ws, title: str, header_row: int, category_col: int, max_row: int, value_col: int, anchor: str) -> None:
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.height = 7.5
    chart.width = 12
    chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chart.firstSliceAng = 270
    chart.add_data(Reference(ws, min_col=value_col, min_row=header_row, max_row=max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=category_col, min_row=header_row + 1, max_row=max_row))
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = False
    chart.dLbls.showPercent = True
    chart.dLbls.showCatName = True
    chart.dLbls.showSerName = False
    chart.dLbls.showLegendKey = False
    chart.dLbls.showBubbleSize = False
    chart.dLbls.showLeaderLines = True
    chart.dLbls.dLblPos = "bestFit"
    chart.dLbls.separator = "\n"
    if chart.legend:
        chart.legend.position = "r"
    ws.add_chart(chart, anchor)


def add_q47_chart(ws, title: str, header_row: int, category_col: int, max_row: int, anchor: str) -> None:
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = title
    chart.height = 7.5
    chart.width = 15
    chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chart.add_data(Reference(ws, min_col=2, max_col=5, min_row=header_row, max_row=max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=category_col, min_row=header_row + 1, max_row=max_row))
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = False
    chart.dLbls.showLegendKey = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.showPercent = False
    chart.dLbls.showBubbleSize = False
    ws.add_chart(chart, anchor)


def format_workbook(wb) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    if cell.number_format in (None, "General"):
                        cell.number_format = "0.00"
                if isinstance(cell.value, str):
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = max(
                ws.column_dimensions[ws.cell(1, col).column_letter].width or 10,
                12,
            )

    institute_ws = wb["就绪度-领域内研究所"]
    institute_ws.column_dimensions["A"].width = 13
    institute_ws.column_dimensions["B"].width = 28
    for column in "CDEFGHIJ":
        institute_ws.column_dimensions[column].width = 14
    institute_ws.column_dimensions["K"].width = 3
    institute_ws.column_dimensions["R"].width = 3
    institute_ws.column_dimensions["S"].width = 16
    institute_ws.column_dimensions["T"].width = 28
    for col_idx in range(21, 27):
        institute_ws.column_dimensions[get_column_letter(col_idx)].width = 14
    institute_ws.column_dimensions["AA"].width = 3
    overall_ws = wb["就绪度-整体"]
    overall_ws.column_dimensions["A"].width = 52
    overall_ws.column_dimensions["B"].width = 16
    overall_ws.column_dimensions["D"].width = 3
    field_ws = wb["就绪度-领域整体"]
    field_ws.column_dimensions["A"].width = 18
    field_ws.column_dimensions["U"].width = 36
    for column in "BCDEFGHIJKLMNOPQ":
        field_ws.column_dimensions[column].width = 12
    for column in ("V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK"):
        field_ws.column_dimensions[column].width = 12
    for sheet_name in ("需求", "就绪度-多选"):
        ws = wb[sheet_name]
        ws.column_dimensions["A"].width = 52
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 8


def build_summary(score_dir: Path, template_path: Path, output_path: Path) -> None:
    if template_path.resolve() == output_path.resolve():
        raise SystemExit(
            f"拒绝覆盖 demo 示例文件：{template_path}\n"
            "请使用 --output 指向一个新文件。"
        )
    if not template_path.is_file():
        raise SystemExit(f"demo 示例文件不存在：{template_path}")

    records = load_records(score_dir)
    shutil.copy2(template_path, output_path)

    workbook = load_workbook(output_path)
    required_sheets = {
        "研究领域分布",
        "就绪度-整体",
        "就绪度-领域整体",
        "就绪度-领域内研究所",
        "就绪度-多选",
        "需求",
    }
    missing = required_sheets - set(workbook.sheetnames)
    if missing:
        raise SystemExit(f"demo 示例文件缺少工作表：{sorted(missing)}")

    populate_field_distribution(workbook, records)
    populate_overall_readiness(workbook, records)
    populate_field_readiness(workbook, records)
    populate_institute_readiness(workbook, records)
    populate_multiselect(workbook, records)
    populate_demand(workbook, records)
    format_workbook(workbook)
    workbook.save(output_path)
    print(f"wrote {output_path}")


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="汇总 score_tables 并生成问卷图表设计工作簿。")
    parser.add_argument("--score-dir", type=Path, default=DEFAULT_SCORE_DIR)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE, help="只读 demo 示例文件")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="生成的汇总工作簿")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    build_summary(args.score_dir, args.template, args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
